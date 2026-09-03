"""Export helpers shared by manual and scheduled reports."""
from __future__ import annotations

import csv
import io
import json
import math
import os
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Mapping, Sequence
from uuid import uuid4

from flask import send_file

from utils.jalali import gregorian_to_jalali
from utils.local_time import local_now_naive


def safe_filename(value: str) -> str:
    value = re.sub(r'[^\w\-.]+', '-', str(value or 'report'), flags=re.UNICODE).strip('-')
    return value[:100] or 'report'


def display_value(value, kind: str = 'text') -> str:
    if value is None:
        return ''
    if kind == 'date' and isinstance(value, (date, datetime)):
        return gregorian_to_jalali(value)
    if kind == 'datetime' and isinstance(value, datetime):
        return f'{gregorian_to_jalali(value)} {value:%H:%M}'
    if kind == 'money':
        try:
            return f'{Decimal(str(value)):,.0f}'
        except Exception:
            return str(value)
    if kind == 'percent':
        try:
            return f'{Decimal(str(value)):,.1f}%'
        except Exception:
            return str(value)
    if isinstance(value, datetime):
        return value.isoformat(timespec='seconds')
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return str(value)


def spreadsheet_safe(value: str) -> str:
    """Prevent user-controlled cells from becoming formulas in Excel/CSV."""
    text = str(value)
    if text.lstrip().startswith(('=', '+', '-', '@', '\t', '\r')):
        return "'" + text
    return text


def selected_columns(result: Mapping, keys: Sequence[str] | None = None) -> list[dict]:
    columns = list(result['columns'])
    if not keys:
        return columns
    allowed = set(keys)
    selected = [column for column in columns if column['key'] in allowed]
    return selected or columns


def tabular_data(result: Mapping, keys: Sequence[str] | None = None):
    columns = selected_columns(result, keys)
    headers = [column['label'] for column in columns]
    rows = []
    for row in result['rows']:
        rendered = []
        for column in columns:
            value = row.get(column['key'])
            kind = column.get('type', 'text')
            text = display_value(value, kind)
            if kind in ('money', 'number', 'percent'):
                try:
                    number = Decimal(str(value))
                    rendered.append(text if number.is_finite() else spreadsheet_safe(text))
                except (InvalidOperation, ValueError, TypeError):
                    rendered.append(spreadsheet_safe(text))
            else:
                rendered.append(spreadsheet_safe(text))
        rows.append(rendered)
    return columns, headers, rows


def csv_bytes(result: Mapping, keys: Sequence[str] | None = None) -> bytes:
    _columns, headers, rows = tabular_data(result, keys)
    output = io.StringIO(newline='')
    writer = csv.writer(output, lineterminator='\n')
    writer.writerow(headers)
    writer.writerows(rows)
    # UTF-8 BOM makes Persian text open correctly in desktop Excel.
    return ('\ufeff' + output.getvalue()).encode('utf-8')


def excel_bytes(result: Mapping, keys: Sequence[str] | None = None) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    columns = selected_columns(result, keys)
    wb = Workbook()
    ws = wb.active
    ws.title = 'گزارش'
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = 'A3'

    title = result['meta']['title']
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(columns)))
    cell = ws.cell(1, 1, title)
    cell.font = Font(name='Tahoma', size=15, bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='0D47A1')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    header_fill = PatternFill('solid', fgColor='E8EEF8')
    border = Border(bottom=Side(style='thin', color='CBD5E1'))
    for index, column in enumerate(columns, 1):
        header = ws.cell(2, index, column['label'])
        header.font = Font(name='Tahoma', size=10, bold=True, color='1E3A8A')
        header.fill = header_fill
        header.alignment = Alignment(horizontal='center', vertical='center')
        header.border = border

    for row_index, row in enumerate(result['rows'], 3):
        for column_index, column in enumerate(columns, 1):
            value = row.get(column['key'])
            kind = column.get('type', 'text')
            cell = ws.cell(row_index, column_index)
            if kind in ('money', 'number', 'percent') and value not in (None, ''):
                try:
                    number = Decimal(str(value))
                    if not number.is_finite():
                        raise ValueError('non-finite numeric cell')
                    cell.value = float(number)
                    cell.number_format = '#,##0.00' if kind != 'percent' else '0.0"%"'
                except (InvalidOperation, ValueError, TypeError):
                    cell.value = spreadsheet_safe(display_value(value, kind))
            elif kind in ('date', 'datetime'):
                cell.value = display_value(value, kind)
            else:
                cell.value = spreadsheet_safe(display_value(value, kind))
            cell.font = Font(name='Tahoma', size=9)
            cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
            cell.border = border

    if result['rows']:
        ws.auto_filter.ref = f'A2:{get_column_letter(len(columns))}{len(result["rows"])+2}'
    for idx, column in enumerate(columns, 1):
        samples = [column['label']] + [display_value(row.get(column['key']), column.get('type', 'text'))
                                       for row in result['rows'][:250]]
        width = min(42, max(11, max((len(str(item)) for item in samples), default=11) + 2))
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.oddFooter.center.text = 'Academy Manager Pro — مرکز گزارش‌ها'
    ws.oddFooter.right.text = 'صفحه &P از &N'
    ws.print_title_rows = '1:2'
    ws.page_setup.orientation = 'landscape' if len(columns) > 6 else 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _json_value(value):
    if isinstance(value, Decimal):
        return float(value) if value.is_finite() else None
    if isinstance(value, float):
        return value if math.isfinite(value) else None
    if isinstance(value, datetime):
        return value.isoformat(timespec='seconds')
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Mapping):
        return {key: _json_value(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_value(item) for item in value]
    return value


def json_bytes(result: Mapping, keys: Sequence[str] | None = None) -> bytes:
    columns = selected_columns(result, keys)
    payload = {
        'report': result['meta']['key'],
        'title': result['meta']['title'],
        'generated_at': result['generated_at'].isoformat(timespec='seconds'),
        'columns': columns,
        'rows': [
            {column['key']: _json_value(row.get(column['key'])) for column in columns}
            for row in result['rows']
        ],
        'kpis': [
            {**item, 'value': _json_value(item.get('value'))}
            for item in result.get('kpis', [])
        ],
        'footers': {key: _json_value(value)
                    for key, value in result.get('footers', {}).items()},
        'chart': _json_value(result.get('chart')),
        'warnings': list(result.get('warnings', [])),
        'filters': (result['filters'].serialisable()
                    if hasattr(result.get('filters'), 'serialisable') else {}),
        'total_rows': result.get('total_rows', len(result['rows'])),
        'exported_rows': len(result['rows']),
    }
    return json.dumps(payload, ensure_ascii=False, indent=2, allow_nan=False).encode('utf-8')


def export_bytes(result: Mapping, export_format: str,
                 keys: Sequence[str] | None = None) -> tuple[bytes, str]:
    export_format = export_format.lower()
    if export_format == 'csv':
        return csv_bytes(result, keys), 'text/csv'
    if export_format in ('xlsx', 'excel'):
        return excel_bytes(result, keys), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    if export_format == 'json':
        return json_bytes(result, keys), 'application/json'
    raise ValueError('فرمت خروجی معتبر نیست')


def export_response(result: Mapping, export_format: str,
                    keys: Sequence[str] | None = None):
    data, mimetype = export_bytes(result, export_format, keys)
    extension = 'xlsx' if export_format in ('xlsx', 'excel') else export_format
    filename = f'{safe_filename(result["meta"]["key"])}-{local_now_naive():%Y%m%d-%H%M}.{extension}'
    return send_file(
        io.BytesIO(data), mimetype=mimetype, as_attachment=True,
        download_name=filename, max_age=0,
    )


def write_export_file(result: Mapping, export_format: str, folder: str | os.PathLike,
                      keys: Sequence[str] | None = None) -> Path:
    data, _mimetype = export_bytes(result, export_format, keys)
    extension = 'xlsx' if export_format in ('xlsx', 'excel') else export_format
    target_dir = Path(folder)
    target_dir.mkdir(parents=True, exist_ok=True)
    target = target_dir / (
        f'{safe_filename(result["meta"]["key"])}-{local_now_naive():%Y%m%d-%H%M%S}-'
        f'{uuid4().hex[:8]}.{extension}'
    )
    target.write_bytes(data)
    return target
