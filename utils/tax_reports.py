"""ابزارهای ساخت خروجی گزارش‌های مالیاتی (Excel / متنی)."""
from __future__ import annotations

import csv
import io

from flask import make_response


def excel_response(filename: str, sheets: dict[str, dict]) -> 'flask.Response':
    """ساخت فایل اکسل چند شیتی.

    sheets = {'نام شیت': {'headers': [...], 'rows': [[...], ...], 'widths': [...]}}
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    workbook.remove(workbook.active)

    header_font = Font(bold=True, color='FFFFFF', name='Tahoma', size=10)
    header_fill = PatternFill('solid', fgColor='0D47A1')
    body_font = Font(name='Tahoma', size=10)

    for sheet_name, data in sheets.items():
        sheet = workbook.create_sheet(title=sheet_name[:31])
        sheet.sheet_view.rightToLeft = True
        headers = data.get('headers', [])
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for row in data.get('rows', []):
            sheet.append(list(row))
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = body_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
        widths = data.get('widths') or [18] * len(headers)
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        sheet.freeze_panes = 'A2'

    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)

    response = make_response(stream.read())
    response.headers['Content-Type'] = \
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def csv_response(filename: str, headers: list[str], rows: list) -> 'flask.Response':
    """خروجی CSV با BOM تا در اکسل فارسی درست باز شود."""
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(headers)
    writer.writerows(rows)

    response = make_response('\ufeff' + buffer.getvalue())
    response.headers['Content-Type'] = 'text/csv; charset=utf-8'
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def text_response(filename: str, content: str) -> 'flask.Response':
    response = make_response(content)
    response.headers['Content-Type'] = 'text/plain; charset=utf-8'
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ═══════════════════════════════════════════════════════════════
#  فایل لیست مالیات حقوق (salary.tax.gov.ir)
# ═══════════════════════════════════════════════════════════════
SALARY_TXT_HEADERS = [
    'ردیف', 'نام', 'نام خانوادگی', 'کد ملی', 'شماره بیمه', 'شغل',
    'حقوق ماهانه', 'مزایای مستمر', 'مزایای غیرمستمر', 'ناخالص',
    'معافیت', 'کسور بیمه', 'درآمد مشمول', 'مالیات', 'خالص پرداختی',
]


def build_salary_tax_txt(rows: list[dict], period: str, employer_code: str = '') -> str:
    """تولید فایل متنی جداشده با «|» برای بارگذاری در سامانه مالیات حقوق.

    ساختار ستون‌ها با هدر SALARY_TXT_HEADERS مطابقت دارد؛ سازمان قالب رسمی خود را
    ممکن است به‌روزرسانی کند، بنابراین خروجی اکسل نیز در کنار آن ارائه می‌شود.
    """
    lines = [f'#PERIOD={period}', f'#EMPLOYER={employer_code}',
             '#' + '|'.join(SALARY_TXT_HEADERS)]
    for index, row in enumerate(rows, start=1):
        lines.append('|'.join(str(value) for value in [
            index,
            row.get('first_name', ''),
            row.get('last_name', ''),
            row.get('national_code', ''),
            row.get('insurance_number', ''),
            row.get('job_title', ''),
            int(row.get('base', 0)),
            int(row.get('continuous', 0)),
            int(row.get('non_continuous', 0)),
            int(row.get('gross', 0)),
            int(row.get('exemption', 0)),
            int(row.get('insurance', 0)),
            int(row.get('taxable', 0)),
            int(row.get('tax', 0)),
            int(row.get('net', 0)),
        ]))
    return '\r\n'.join(lines) + '\r\n'


# ═══════════════════════════════════════════════════════════════
#  گزارش معاملات فصلی (ماده ۱۶۹ / TTMS)
# ═══════════════════════════════════════════════════════════════
TTMS_SALE_HEADERS = [
    'نوع معامله', 'شماره صورتحساب', 'تاریخ', 'نوع شخص', 'نام طرف معامله',
    'شناسه ملی/کد ملی', 'کد اقتصادی', 'کد پستی', 'نشانی',
    'شرح کالا/خدمت', 'مبلغ کل', 'تخفیف', 'مبلغ مشمول', 'مبلغ معاف',
    'ارزش افزوده', 'جمع کل', 'شماره منحصر به فرد مالیاتی',
]


def quarter_of(jalali_month: int) -> int:
    """شماره فصل شمسی (۱ تا ۴) از شماره ماه."""
    return (max(1, min(12, int(jalali_month))) - 1) // 3 + 1


QUARTER_NAMES = {1: 'بهار', 2: 'تابستان', 3: 'پاییز', 4: 'زمستان'}
QUARTER_MONTHS = {1: (1, 3), 2: (4, 6), 3: (7, 9), 4: (10, 12)}


def quarter_date_range(year: int, quarter: int):
    """بازه میلادی متناظر با یک فصل شمسی."""
    from datetime import timedelta

    import jdatetime

    start_month, end_month = QUARTER_MONTHS[quarter]
    start = jdatetime.date(year, start_month, 1).togregorian()
    if end_month == 12:
        next_start = jdatetime.date(year + 1, 1, 1).togregorian()
    else:
        next_start = jdatetime.date(year, end_month + 1, 1).togregorian()
    return start, next_start - timedelta(days=1)
